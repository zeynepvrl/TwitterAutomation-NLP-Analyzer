from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from textblob import TextBlob
import time
import re
from datetime import datetime
import winsound
import openpyxl
from openpyxl.utils import get_column_letter
import os

# WebDriver'ı başlatın
service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service)


def check(username):
    driver.get(f"https://twitter.com/{username}")
    time.sleep(3)

    # Katılma tarihini kontrol etmek için span'ların XPath'lerini belirleyin
    span_xpaths = [
        "//*[@id='react-root']/div/div/div/main/div/div/div/div/div/div/div/div/div/div/div/div/div/span[1]",
        "//*[@id='react-root']/div/div/div/main/div/div/div/div/div/div/div/div/div/div/div/div/span[2]",
        "//*[@id='react-root']/div/div/div/main/div/div/div/div/div/div/div/div/div/div/div/div/span[3]",
        "//*[@id='react-root']/div/div/div[2]/main/div/div/div/div/div/div[3]/div/div/div/div[2]/div[3]/div/span/span",
        "//*[@id='react-root']/div/div/div[2]/main/div/div/div/div/div/div[3]/div/div/div/div/div[3]/div/span/span",
        "//*[@id='react-root']/div/div/div[2]/main/div/div/div/div/div/div[3]/div/div/div/div/div[4]/div/span/span",
        "//*[@id='react-root']/div/div/div[2]/main/div/div/div/div/div/div[3]/div/div/div[1]/div/div[4]/div/span/span",
        "//*[@id='react-root']/div/div/div[2]/main/div/div/div/div/div/div[3]/div/div/div[1]/div/div[4]/div/span[4]/span"
    ]

    join_year = None

    for xpath in span_xpaths:
        try:
            span_text = driver.find_element(By.XPATH, xpath).text
            if 'Joined' in span_text:
                join_year = span_text[-4:]  # Son 4 karakteri (yıl) al
                break  # Eğer bulduysak döngüden çık
        except:
            continue  # Eğer bir span bulunamazsa, bir sonrakine geç


    following_xpaths = [
        "//*[@id='react-root']/div/div/div[2]/main/div/div/div/div/div/div[3]/div/div/div/div/div[4]/div[1]/a",
        "//*[@id='react-root']/div/div/div[2]/main/div/div/div/div/div/div[3]/div/div/div/div[2]/div[5]/div[1]/a",
        "//*[@id='react-root']/div/div/div[2]/main/div/div/div/div/div/div[3]/div/div/div/div/div[5]/div[1]/a",
        "//*[@id='react-root']/div/div/div[2]/main/div/div/div/div/div/div[3]/div/div/div[1]/div[2]/div[5]/div[1]/button",
        "//*[@id='react-root']/div/div/div[2]/main/div/div/div/div/div/div[3]/div/div/div[1]/div[2]/div[4]/div[1]/button",
        "//*[@id='react-root']/div/div/div[2]/main/div/div/div/div/div/div[3]/div/div/div[1]/div/div[5]/div[1]/button",
        "//*[@id='react-root']/div/div/div[2]/main/div/div/div/div/div/div[3]/div/div/div[1]/div[2]/div[5]/div[1]/a",
        "//*[@id='react-root']/div/div/div[2]/main/div/div/div/div/div/div[3]/div/div/div[1]/div/div[5]/div[1]/a",
        "//*[@id='react-root']/div/div/div[2]/main/div/div/div/div/div/div[3]/div/div/div/div[2]/div[4]/div[1]/a",
        "//*[@id='react-root']/div/div/div[2]/main/div/div/div/div/div/div[3]/div/div/div[1]/div/div[4]/div[1]/button",
        "//*[@id='react-root']/div/div/div[2]/main/div/div/div/div/div/div[3]/div/div/div/div/div[6]/div[1]/a"
    ]
    following = None
    for xpath in following_xpaths:
        try:
            span_text = driver.find_element(By.XPATH, xpath).text
            if 'Following' in  span_text:
                following=span_text.split(" ")[0]         
                break  # Eğer bulduysak döngüden çık
        except:
            continue  # Eğer bir span bulunamazsa, bir sonrakine geç


    follower_xpaths = [
        "//*[@id='react-root']/div/div/div[2]/main/div/div/div/div/div/div[3]/div/div/div[1]/div[2]/div[5]/div[2]/a",
        "//*[@id='react-root']/div/div/div[2]/main/div/div/div/div/div/div[3]/div/div/div/div/div[5]/div[2]/a",
        "//*[@id='react-root']/div/div/div[2]/main/div/div/div/div/div/div[3]/div/div/div[1]/div[2]/div[4]/div[2]/button",
        "//*[@id='react-root']/div/div/div[2]/main/div/div/div/div/div/div[3]/div/div/div[1]/div[2]/div[4]/div[2]/a",
        "//*[@id='react-root']/div/div/div[2]/main/div/div/div/div/div/div[3]/div/div/div/div/div[4]/div[2]/a",
        "//*[@id='react-root']/div/div/div[2]/main/div/div/div/div/div/div[3]/div/div/div[1]/div/div[4]/div[2]/a",
        "//*[@id='react-root']/div/div/div[2]/main/div/div/div/div/div/div[3]/div/div/div[1]/div/div[5]/div[2]/a",
        "//*[@id='react-root']/div/div/div[2]/main/div/div/div/div/div/div[3]/div/div/div/div[2]/div[4]/div[2]/a",
        "//*[@id='react-root']/div/div/div[2]/main/div/div/div/div/div/div[3]/div/div/div/div[2]/div[5]/div[2]/a",
        "//*[@id='react-root']/div/div/div[2]/main/div/div/div/div/div/div[3]/div/div/div[1]/div/div[5]/div[2]/button",
        "//*[@id='react-root']/div/div/div[2]/main/div/div/div/div/div/div[3]/div/div/div[1]/div/div[4]/div[2]/button",
        "//*[@id='react-root']/div/div/div[2]/main/div/div/div/div/div/div[3]/div/div/div/div/div[6]/div[2]/a"
    ]

    follower = None
    for xpath in follower_xpaths:
        try:
            span_text = driver.find_element(By.XPATH, xpath).text
            if 'Followers' in  span_text or 'Follower' in span_text:
                follower=span_text.split(" ")[0]
                break  # Eğer bulduysak döngüden çık
        except:
            continue  # Eğer bir span bulunamazsa, bir sonrakine geç

    try:
        span_text = driver.find_element(By.XPATH, "//*[@id='react-root']/div/div/div[2]/main/div/div/div/div/div/div[1]/div[1]/div/div/div/div/div/div[2]/div/div").text
        postnumber=span_text.split(" ")[0]
    except:
        postnumber=None

    if follower==None or following==None or postnumber==None or join_year==None :
        print("None")
        with open('NONE.txt', 'a' ) as file:
            file.write(f"{username}\n")
        winsound.Beep(1000, 500)
        return False


    if "K" in follower:
        follower=follower.replace(".", "")
        follower=follower.replace("K", "")
        follower=int(follower)*100
    elif "," in follower:
        follower =follower.replace(",", "")  # Virgülü kaldır

    if "K" in following:
        following=following.replace(".", "")
        following=following.replace("K", "")
        following=int(following)*1000
    elif "," in following:
        following =following.replace(",", "")  # Virgülü kaldır
    
    if "K" in postnumber:
        postnumber=postnumber.replace(".", "")
        postnumber=int(postnumber.replace("K", "000"))
    elif "," in postnumber:
        postnumber=postnumber.replace(",", "")



    elements = driver.find_elements(By.CLASS_NAME, 'css-146c3p1.r-8akbws.r-krxsd3.r-dnmrzs.r-1udh08x.r-bcqeeo.r-1ttztb7.r-qvutc0.r-37j5jr.r-a023e6.r-rjixqe.r-16dba41.r-bnwqim')

    for element in elements:
        print(element.text)




    file_path = "output.xlsx"  # Kaydedilecek dosya yolu
    headers = ["Username", "Followers", "Following", "Posts", "Join Year"]
    values = [username, follower, following, postnumber, join_year]

    # update_excel(file_path, headers, values)

    

    
        
        
       

    

# Excel işlemlerini gerçekleştirme
def update_excel(file_path, headers, values):
    # Eğer dosya yoksa yeni bir dosya oluştur
    if not os.path.exists(file_path):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(headers)  # Başlıkları ekle
        workbook.save(file_path)

    # Var olan Excel dosyasını aç
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Mevcut başlıkları kontrol et
    current_headers = [cell.value for cell in sheet[1]]  # İlk satırdaki başlıklar
    for header in headers:
        if header not in current_headers:
            current_headers.append(header)  # Eksik başlıkları listeye ekle

    # Başlıkları güncelle
    for col_num, header in enumerate(current_headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

    # Yeni satır ekle
    new_row = [None] * len(current_headers)  # Varsayılan boş değerler
    for header, value in zip(headers, values):
        if header in current_headers:
            col_index = current_headers.index(header) + 1  # Sütun numarası (0 tabanlı değil)
            new_row[col_index - 1] = value

    # Yeni veriyi sayfaya ekle
    sheet.append(new_row)

    # Değişiklikleri kaydet
    workbook.save(file_path)
    print(f"{file_path} güncellendi.")




try:
    # Twitter'a gidin ve giriş yapın
    driver.get('https://twitter.com/')

    choose_login = WebDriverWait(driver, 1000).until(
        EC.element_to_be_clickable((By.XPATH, "//*[@data-testid='loginButton']"))
    )
    driver.execute_script("arguments[0].click();", choose_login)

    username = WebDriverWait(driver, 100).until(
        EC.element_to_be_clickable((By.XPATH, "//input[@name='text']"))
    )
    enter_user = input("Enter phone, username, or email: ")
    username.clear()
    username.send_keys(enter_user)
    username.send_keys(Keys.RETURN)

    password = WebDriverWait(driver, 100).until(
        EC.element_to_be_clickable((By.XPATH, "//input[@name='password']"))
    )
    enter_pass = input("Password: ")
    password.clear()
    password.send_keys(enter_pass)
    password.send_keys(Keys.RETURN)

    followers_section = WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.XPATH, "//*[@data-testid='primaryColumn']//section"))
    )
    

    start_line=0
    end_line=20

    all_followers=set()

    with open('50BINKISI.txt' , 'r' ) as file:
        for current_line, line in enumerate(file, start=1):
            if start_line<= current_line<=end_line:
                all_followers.add(line.strip())
    print(f"Toplem {len(all_followers)} follower çekildi.")
    
    
    count=0
    #Her bir takipçiyi analiz et
    for username in all_followers:
        print(f"Analyzing {username}...")
        print(count)
        count+=1
        #if check(username):
        check(username)
        """ with open('BOTLAR.txt' , 'a' ) as file:
                file.write(f"{username}\n") """
        
        time.sleep(6)
            # Kullanıcıyı engelle (Bu kısmı ihtiyacınıza göre doldurabilirsiniz)
            #more_options = WebDriverWait(driver, 10).until(
            #     EC.element_to_be_clickable((By.XPATH, "//*[@id='react-root']/div/div/div/main/div/div/div/div/div/div/div/div/div/div/div/div/button"))
            # )                                 
            #more_options.click()
            # block_option = WebDriverWait(driver, 10).until(
            #     EC.element_to_be_clickable((By.XPATH, "//*[@id='layers']/div[2]/div/div/div/div[2]/div/div[3]/div/div/div/div[4]/div[2]/div/span"))
            # )
            # block_option.click()
            # confirm_block = WebDriverWait(driver, 10).until(
            #     EC.element_to_be_clickable((By.XPATH, "/html/body/div/div/div/div/div/div/div/div/div/div/div/div/div/button"))
            # )
            # confirm_block.click()

finally:
    #print(f"Toplam {len(all_followers)} takipçi alındı.")
    winsound.Beep(1000, 800)  # 1000 Hz frekansında 500 milisaniye boyunca bip sesi
    driver.quit()


